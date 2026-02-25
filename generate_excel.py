#!/usr/bin/env python3
"""
Script pour générer Sprint Backlog Excel avec cellules fusionnées
Usage: python generate_excel.py
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("🚀 Génération du Sprint Backlog Excel...")

# Créer workbook
wb = Workbook()
ws = wb.active
ws.title = "Sprint Backlog"

# En-têtes
headers = ['ID US', 'User Story (US)', 'ID Tâche', 'Tâches effectuées', 'Fichiers créés/modifiés', 'Estimation', 'Responsable', 'Sprint', 'Statut']
ws.append(headers)

# Style en-tête
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Lire CSV
csv_file = 'SPRINT_BACKLOG_REEL.csv'
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # Skip header
    data = list(reader)

# Ajouter données
for row in data:
    ws.append(row)

print(f"✅ {len(data)} tâches chargées")

# Fusionner cellules par User Story
merge_ranges = []
current_us = ''
start_row = 2
count = 0

for i, row in enumerate(data):
    us_id = row[0]
    
    if us_id != current_us and current_us != '':
        end_row = start_row + count - 1
        if count > 1:
            merge_ranges.append({
                'A': f'A{start_row}:A{end_row}',
                'B': f'B{start_row}:B{end_row}'
            })
        start_row = i + 2
        count = 1
        current_us = us_id
    elif us_id == current_us:
        count += 1
    else:
        current_us = us_id
        count = 1

# Dernière plage
end_row = start_row + count - 1
if count > 1:
    merge_ranges.append({
        'A': f'A{start_row}:A{end_row}',
        'B': f'B{start_row}:B{end_row}'
    })

# Appliquer fusions
for ranges in merge_ranges:
    ws.merge_cells(ranges['A'])
    ws.merge_cells(ranges['B'])

print(f"✅ {len(merge_ranges)} User Stories fusionnées")

# Largeur colonnes
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12

# Hauteur lignes
ws.row_dimensions[1].height = 25

# Styles pour toutes les cellules
center_alignment = Alignment(horizontal="center", vertical="center")
wrap_alignment = Alignment(vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=9):
    for cell in row:
        cell.border = thin_border
        
    # Alignement
    row[0].alignment = center_alignment  # ID US
    row[1].alignment = wrap_alignment     # User Story
    row[2].alignment = center_alignment  # ID Tâche
    row[3].alignment = wrap_alignment     # Tâches
    row[4].alignment = wrap_alignment     # Fichiers
    row[5].alignment = center_alignment  # Estimation
    row[6].alignment = center_alignment  # Responsable
    row[7].alignment = center_alignment  # Sprint
    row[8].alignment = center_alignment  # Statut

# Sauvegarder
wb.save('SPRINT_BACKLOG_REEL.xlsx')

print("✅ Fichier Excel créé avec succès!")
print("📁 Fichier: SPRINT_BACKLOG_REEL.xlsx")
print("📊 Cellules fusionnées pour éviter les redondances!")
print("🎉 Terminé!")
