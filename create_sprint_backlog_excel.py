#!/usr/bin/env python3
"""
Script pour créer SPRINT_BACKLOG_REEL.xlsx avec cellules fusionnées
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

# Lire le CSV
rows = []
with open('SPRINT_BACKLOG_REEL.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    rows = list(reader)

# Créer le workbook
wb = Workbook()
ws = wb.active
ws.title = "Sprint Backlog"

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Couleurs par sprint
sprint_colors = {
    "Sprint 1": "D6EAF8",  # Bleu clair
    "Sprint 2": "D5F4E6",  # Vert clair
    "Sprint 3": "FCF3CF"   # Orange clair
}

# Écrire les en-têtes
headers = rows[0]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Grouper les lignes par US
us_groups = defaultdict(list)
for row in rows[1:]:
    us_id = row[0]
    us_groups[us_id].append(row)

# Fonction de tri numérique pour US-X.Y
def sort_us_key(us_id):
    """Trie US-1.1, US-1.2, ..., US-1.10, US-1.11 correctement"""
    try:
        # Extraire les nombres de "US-1.10" -> [1, 10]
        parts = us_id.replace('US-', '').split('.')
        return [int(p) for p in parts]
    except:
        return [0, 0]

# Écrire les données avec fusion
current_row = 2
for us_id in sorted(us_groups.keys(), key=sort_us_key):
    tasks = us_groups[us_id]
    start_row = current_row
    
    for task in tasks:
        for col_idx, value in enumerate(task, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Couleur selon sprint
            sprint = task[7] if len(task) > 7 else ""
            if sprint in sprint_colors:
                cell.fill = PatternFill(start_color=sprint_colors[sprint], 
                                       end_color=sprint_colors[sprint], 
                                       fill_type="solid")
        
        current_row += 1
    
    # Fusionner les cellules pour ID US et User Story
    if len(tasks) > 1:
        ws.merge_cells(start_row=start_row, start_column=1, 
                      end_row=current_row-1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, 
                      end_row=current_row-1, end_column=2)

        # Centrer le texte dans les cellules fusionnées
        ws.cell(row=start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.cell(row=start_row, column=2).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

# Ajuster la largeur des colonnes
column_widths = {
    1: 8,   # ID US
    2: 50,  # User Story
    3: 8,   # ID Tâche
    4: 60,  # Tâches effectuées
    5: 50,  # Fichiers
    6: 12,  # Estimation
    7: 15,  # Responsable
    8: 10,  # Sprint
    9: 10   # Statut
}

for col, width in column_widths.items():
    ws.column_dimensions[chr(64 + col)].width = width

# Figer la première ligne
ws.freeze_panes = "A2"

# Sauvegarder
wb.save('SPRINT_BACKLOG_REEL.xlsx')
print("✅ Fichier SPRINT_BACKLOG_REEL.xlsx créé avec succès!")
print(f"📊 Total: {current_row - 2} tâches")
print(f"📋 User Stories: {len(us_groups)}")
